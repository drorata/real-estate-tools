from io import BytesIO
from tempfile import NamedTemporaryFile

import pandas as pd
import plotly.express as px
import streamlit as st
from loguru import logger
from openpyxl import Workbook

from real_estate_tools import flip_calculator_components as fcc
from real_estate_tools.utils import summarize_flip_calc

st.set_page_config(layout="wide")

logger.info("Starting a flip calculation cycle")

"""
# Flip Calculator

Fill in all the details and checkout the graph at the bottom to see what is the best
purchase price of your property
"""
general_col, selling_col, purchase_col, holding_costs_col, rehab_col = st.columns(5)


###
# General section
###
general_col.write("## General factors")
arv = general_col.number_input(
    label="Provide estimated ARV ($)",
    value=240000,
    step=1000,
    help="This will be used to derive the selling price.",
)
attractivity_factor = (
    general_col.number_input(
        label="Provide the attractivity factor (%)",
        value=2,
        min_value=0,
        max_value=100,
    )
    / 100
)
listing_price = general_col.number_input(
    label="What is the listing price?",
    value=170000,
    step=1000,
)
min_purchase_price = general_col.number_input(
    label="What's the minimal purchase price ($)?",
    min_value=0,
    max_value=int(arv * (1 - attractivity_factor)),
    step=1000,
    value=int(listing_price * 0.7),
)
estimated_holding_period = general_col.number_input(
    label="Number of months expected for flip", min_value=1, value=3
)

###
# Selling section
###
selling_col.write(
    "## Selling factors\n"
    ":blue[Provide estimations of the different factors to impact the **selling**"
    " costs]"
)
selling_factors = fcc.sell_or_purchase_factors(selling_col, mode="sell")

###
# Purchase section
###
purchase_col.write(
    "## Purchase factors\n"
    ":blue[Provide estimations of the different factors to impact the **purchase**"
    " costs]"
)
purchase_factors = fcc.sell_or_purchase_factors(purchase_col, mode="purchase")


###
# Holding costs section
###
holding_costs_col.write("## Monthly holding costs\nIndicate expected monthly costs")
monthly_costs = pd.Series(
    {
        "taxes_per_month": holding_costs_col.number_input(
            label="Tax ($)", min_value=0.0, step=10.0
        ),
        "insurance_per_month": holding_costs_col.number_input(
            label="Insurance ($)", min_value=0.0, step=10.0
        ),
        "utilities_gas_per_month": holding_costs_col.number_input(
            label="Gas ($)", min_value=0.0, step=10.0
        ),
        "utilities_electricity_per_month": holding_costs_col.number_input(
            label="Electricity ($)", min_value=0.0, step=10.0
        ),
        "utilities_water_sewer_per_month": holding_costs_col.number_input(
            label="Water and sewer ($)", min_value=0.0, step=10.0
        ),
        "trash_per_month": holding_costs_col.number_input(
            label="Trash ($)", min_value=0.0, step=10.0
        ),
    },
    name="monthly_costs",
)
holding_costs = pd.DataFrame({
    "total": monthly_costs * estimated_holding_period,
    "monthly": monthly_costs,
})

###
# Rehab costs section
###
rehab_col.write("## Rehab costs\nProvide estimation of the rehab total costs")
total_rehab = rehab_col.number_input(
    label="Rehab costs ($)", min_value=0, step=500, value=int(arv * 0.05)
)


"""## Summary"""
result = pd.DataFrame({"purchase_price": range(min_purchase_price, arv, 1000)})
result["purchase_costs"] = result["purchase_price"].apply(
    lambda x: purchase_factors.get_cost(x)
)
result["selling_costs"] = selling_factors.get_cost(arv * (1 - attractivity_factor))
result["holding_costs"] = holding_costs["total"].sum()
result["rehab_costs"] = total_rehab
result["total_expenses"] = result.sum(axis=1)
result["total_income"] = arv * (1 - attractivity_factor)
result["ROI"] = 100 * (result["total_income"] / result["total_expenses"] - 1)

idx_of_20_pct_ROI = (result["ROI"] - 20).abs().idxmin()
fig = px.line(
    result, x="purchase_price", y="ROI", labels={"purchase_price": "Purchase price"}
)
fig.add_vline(
    x=result.iloc[idx_of_20_pct_ROI]["purchase_price"],
    annotation_text="20% ROI",
    line_color="green",
    line_dash="dot",
)
fig.add_vline(
    x=listing_price,
    annotation_text="Listing price",
    line_color="red",
    line_dash="dot",
)
fig.update_layout(
    annotations=[
        {**a, **{"textangle": 90}} for a in fig.to_dict()["layout"]["annotations"]
    ]
)
st.plotly_chart(fig)


workbook = Workbook()
logger.debug(
    f"New workbook has the following sheets: {workbook.sheetnames}. These sheets will"
    " be deleted"
)
for sheet in workbook.sheetnames:
    workbook.remove(workbook[sheet])

output_general = pd.DataFrame(
    [[
        arv,
        attractivity_factor,
        listing_price,
        min_purchase_price,
        estimated_holding_period,
    ]],
    columns=[
        "ARV",
        "Attractivity factor",
        "Listing price",
        "Min purchase price",
        "Estimated holding period",
    ],
)
output_purchase_sell = pd.DataFrame(
    [purchase_factors.model_dump(), selling_factors.model_dump()]
).rename({0: "Purchase", 1: "Selling"}, axis=0)
workbook = summarize_flip_calc(
    workbook=workbook,
    general=output_general,
    purchase_sell_cost_factors=output_purchase_sell,
    holding_costs=holding_costs,
    rehab_costs=total_rehab,
)

with NamedTemporaryFile() as tmp:
    logger.debug(f"Obtain the following temp file {tmp.name}")
    workbook.save(tmp.name)
    data = BytesIO(tmp.read())

show_result_df = st.toggle(label="Show resulting table", value=False)
if show_result_df:
    st.dataframe(result, height=150)

st.download_button(
    label="Download summary (xlsx)",
    data=data,
    mime="application/octet-stream",
    file_name="Flip calculator summary.xlsx",
)
