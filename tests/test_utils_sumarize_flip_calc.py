import pandas as pd
from openpyxl import Workbook

from real_estate_tools.utils import summarize_flip_calc


def test_default():
    workbook = Workbook()

    assert "Flip Calculator" not in workbook.sheetnames

    workbook = summarize_flip_calc(
        workbook=workbook,
        general=pd.DataFrame([[1, 2, 3, 4]], columns=["a", "b", "c", "d"]),
        purchase_sell_cost_factors=pd.DataFrame(
            [[1.1, 2.2, 3.3, 4.4]], columns=["a", "b", "c", "d"]
        ),
        holding_costs=pd.DataFrame(
            [[1.11, 2.22, 3.33, 4.44]], columns=["a", "b", "c", "d"]
        ),
        rehab_costs=1.2,
    )

    assert "Flip Calculator" in workbook.sheetnames
